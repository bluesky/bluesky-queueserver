# -------------------------------------------------------------------------------------
# This module include beamline-specific code. It will be moved into a separate package.
# -------------------------------------------------------------------------------------
import re
from openpyxl import load_workbook
import pandas as pd


def isfloat(value):
    try:
        float(value)
        return True
    except ValueError:
        return False


PERIODIC_TABLE = "\
H                                                                                            He \
Li Be                                                                          B  C  N  O  F Ne \
Na Mg                                                                         Al Si  P  S Cl Ar \
K  Ca                                           Sc Ti  V Cr Mn Fe Co Ni Cu Zn Ga Ge As Se Br Kr \
Rb Sr                                            Y Zr Nb Mo Tc Ru Rh Pd Ag Cd In Sn Sb Te  I Xe \
Cs Ba La Ce Pr Nd Pm Sm Eu Gd Tb Dy Ho Er Tm Yb Lu Hf Ta W  Re Os Ir Pt Au Hg Tl Pb Bi Po At Rn \
Fr Ra Ac Th Pa U  Np Pu Am Cm Bk Cf Es Fm Md No Lr Rf Ha Sg Bh Hs Mt Ds Rg Cn Nh Fl Mc Lv Ts Og"


def sanitize_step_scan_parameters(bounds, steps, times):
    """Attempt to identify and flag/correct some common scan parameter mistakes."""
    problem = False
    text = ""

    # Bounds is one longer than steps/times, length of steps = length of times
    if (len(bounds) - len(steps)) != 1:
        text += "bounds must have one more item than steps"
        text += "bounds = %s" % " ".join(map(str, bounds))
        text += "steps = %s" % " ".join(map(str, steps))
        problem = True
    if (len(bounds) - len(times)) != 1:
        text += "bounds must have one more item than times"
        text += "bounds = %s" % " ".join(map(str, bounds))
        text += "times = %s\n" % " ".join(map(str, times))
        problem = True

    # Tests of boundary values
    for b in bounds:
        if not isfloat(b) and b[-1:].lower() == "k":
            if not isfloat(b[:-1]):
                text += "%s is not a valid scan boundary value" % b
                problem = True
        elif not isfloat(b):
            text += "%s is not a valid scan boundary value" % b
            problem = True

        if not isfloat(b) and b[:1] == "-" and b[-1:].lower() == "k":
            text += "Negative bounds must be energy-valued, not k-valued (%s)" % b
            problem = True

    # Tests of step size values #
    for s in steps:
        if not isfloat(s) and s[-1:].lower() == "k":
            if not isfloat(s[:-1]):
                text += "%s is not a valid scan step size value" % s
                problem = True
            elif float(s[:-1]) < 0:
                text += "Step sizes cannot be negative (%s)" % s
                problem = True
        elif not isfloat(s):
            text += "%s is not a valid scan step size value" % s
            problem = True

        if isfloat(s) and float(s) < 0:
            text += "Step sizes cannot be negative (%s)" % s
            problem = True
        elif isfloat(s) and float(s) <= 0.09:
            text += "%s is a very small step size!" % s
        elif not isfloat(s) and s[-1:].lower() == "k" and isfloat(s[-1:]) and float(s[:-1]) < 0.01:
            text += "%s is a very small step size!" % s

    # tests of integration time values
    for t in times:
        if not isfloat(t) and t[-1:].lower() == "k":
            if not isfloat(t[:-1]):
                text += "%s is not a valid integration time value" % t
                problem = True
            elif float(t[:-1]) < 0:
                text += "Integration times cannot be negative (%s)" % t
                problem = True
        elif not isfloat(t):
            text += "%s is not a valid integration time value" % t
            problem = True

        if isfloat(t) and float(t) < 0:
            text += "Integration times cannot be negative (%s)" % t
            problem = True
        elif isfloat(t) and float(t) <= 0.1:
            text += "%s is a very short integration time!" % t
        elif not isfloat(t) and t[-1:].lower() == "k" and isfloat(t[-1:]) and float(t[:-1]) < 0.05:
            text += "%s is a very short integration time!" % t

    if text:
        text += "see " + "https://nsls-ii-bmm.github.io/BeamlineManual/xafs.html#scan-regions"

    return problem, text


class BMMMacroBuilder:
    """A base class for parsing specially constructed spreadsheets and
    generating the corresponding BlueSky plan.

    attributes
    ----------
    basename : str
       basename of the spreadsheet
    folder : str
       folder containing spreadsheet, usually same as BMMuser.folder
    joiner : str
       string used to construct filenames [-] (_ is a also a good choice)
    source : str
       fully resolved path to spreadsheet
    wb : openpyxl workbook object
       workbook created from spreadsheet
    ws : openpyxl worksheet object
       main sheet of spreadsheet
    measurements : list
       list of disctionaries, one for each row of the spreadsheet
    ini : str
       fully resolved path to INI file
    macro : str
       fully resolved path to plan file
    tab : str
       string used to pythonically format the plan file
    do_first_change : bool
       True is need to begin with a change_edge()
    has_e0_column : bool
       True is this is a very old wheel spreadsheet
    offset : int
       1 if this is a very old wheel spreadsheet
    verbose : bool
       True for more comment lines in  the plan
    totaltime : float
       estimate for the run time of the plan
    deltatime : float
       estimated uncertainty in the total time estimate
    instrument : str
       "sample wheel" or "glancing angle stage"

    Required method
    ---------------
    _gen_plan_list
       generate the text of the BlueSky plan
    get_keywords
       instructions for parsing spreadsheet columns into keywords

    """

    def __init__(self, *, user_name):
        self.user_name = user_name

        # Generated plan list (returned by the function)
        self.plan_list = []
        # Default parameters that are included with each 'xafs' plan
        self.default_xafs_parameters = []

        self.basename = None
        self.folder = None
        self.joiner = "-"

        self.source = None
        self.wb = None
        self.ws = None
        self.measurements = list()
        self.ini = None
        self.macro = None

        self.tab = " " * 8
        self.do_first_change = False
        self.has_e0_column = False
        self.offset = 0
        self.verbose = False

        self.totaltime = 0
        self.deltatime = 0

        self.tmpl = None
        self.instrument = None

        self.experiment = ("default", "slot", "focus", "measure", "spin", "angle", "method")
        self.flags = ("snapshots", "htmlpage", "usbstick", "bothways", "channelcut", "ththth")
        self.motors = ("samplex", "sampley", "samplep", "slitwidth", "detectorx")
        self.science_metadata = ("url", "doi", "cif")

        self.wini = None
        self.close_shutters = None
        self.append_element = None

    def clear_plan_list(self):
        self.plan_list = []

    def add_plan(self, *, name, args=None, kwargs=None):
        args = args or []
        kwargs = kwargs or {}
        self.plan_list.append({"name": name, "args": args, "kwargs": kwargs})

    def process_spreadsheet(self, *, spreadsheet_file, energy=False):
        """Convert a wheel macro spreadsheet to a BlueSky plan.

        Examples
        --------
        To create a macro from a spreadsheet called "MySamples.xlsx"

        >>> xlsx('MySamples')

        To specify a change_edge() command at the beginning of the macro:

        >>> xlsx('MySamples', energy=True)
        """
        self.wb = load_workbook(spreadsheet_file, read_only=True)
        self.ws = self.wb.active
        # sh_names = self.wb.sheetnames
        # self.ws = self.wb[sh_names[0]]

        self.measurements = list()
        if energy is True:
            self.do_first_change = True

        if self.ws["H5"].value.lower() == "e0":  # accommodate older xlsx files which have e0 values in column H
            self.has_e0_column = True

        self.do_first_change = self.truefalse(self.ws["G2"].value)
        self.close_shutters = self.truefalse(self.ws["J2"].value)
        self.append_element = str(self.ws["L2"].value)

        self.instrument = str(self.ws["B1"].value).lower()

        isok, explanation = self.read_spreadsheet()
        if isok is False:
            raise RuntimeError(f"Error occurred while parsing spreadsheet: {explanation}")

        self.gen_plan_list()

        return self.plan_list

    def truefalse(self, value):
        """Interpret certain strings from the spreadsheet as True/False"""
        if value is None:
            return True  # self.measurements[0]['measure']
        if str(value).lower() == "=true()":
            return True
        elif str(value).lower() == "true":
            return True
        elif str(value).lower() == "yes":
            return True
        else:
            return False

    def ini_sanity(self, default):
        """
        Sanity checks for the default line from the spreadsheet.

        1. experimenters is a string (BMMuser.name)
        2. sample, prep, and comment are not empty strings (set to '...')
        3. nscans is an integer (set to 1)
        4. start is an integer or "next"
        5. mode is string (set to 'transmission')
        6. element is an element (bail)
        7. edge is k, l1, l2, or l3 (bail)

        To do:
          * booleans are interpretable as booleans
          * focused is focused or unfocused
          * bounds, steps, times are sensible
          * x, y, slits are floats and sensible for the respective ranges of motion
        """
        message = ""
        unrecoverable = False

        if "mode" not in default:
            if "glancing angle" in self.instrument:
                default["mode"] = "xs"
            else:
                default["mode"] = "transmission"

        if default["filename"] is None or str(default["filename"]).strip() == "":
            default["filename"] = "filename"

        if default["experimenters"] is None or str(default["experimenters"]).strip() == "":
            default["experimenters"] = self.user_name

        defaultdefaults = {"bounds": "-200  -30  -10 15.5  570", "steps": "10  2  0.25  0.05k", "times": "1 1 1 1"}
        for k in ("bounds", "steps", "times"):
            if default[k] is None or str(default[k]).strip() == "":
                default[k] = defaultdefaults[k]

        for k in ("sample", "prep", "comment"):
            if default[k] is None or str(default[k]).strip() == "":
                default[k] = "..."
            if "%" in default[k]:
                default[k] = default[k].replace("%", "%%")

        try:
            default["nscans"] = int(default["nscans"])
        except Exception:
            default["nscans"] = 1

        try:
            default["start"] = int(default["start"])
        except Exception:
            default["start"] = "next"

        # if default['mode'] is None or str(default['mode']).strip() == '':
        #    default['mode'] = 'transmission'

        if str(default["element"]).capitalize() not in re.split(r"\s+", PERIODIC_TABLE):  # see 06-periodic table
            message += "\nDefault entry for element is not recognized."
            unrecoverable = True

        if str(default["edge"]).lower() not in ("k", "l1", "l2", "l3"):
            message += "\nDefault entry for edge is not recognized."
            unrecoverable = True

        # try:
        #     default['e0'] = float(default['e0'])
        # except:
        #     default['e0'] = edge_energy(default['element'], default['edge'])

        if unrecoverable:
            raise RuntimeError(message)

        return default

    def read_spreadsheet(self):
        """Slurp up the content of the spreadsheet and write the default control file"""
        print("Reading spreadsheet: %s" % self.source)
        count = 0
        self.offset = 0
        isok, explanation = True, ""
        if self.has_e0_column:  # deal with older xlsx that have e0 in column H
            self.offset = 1

        for row in self.ws.rows:
            count += 1
            if count < 6:
                continue
            defaultline = False
            if count == 6:
                defaultline = True
            if count > 200:
                break
            self.measurements.append(self.get_keywords(row, defaultline))

            # check that scan parameters make sense
            if type(self.measurements[-1]["bounds"]) is str:
                b = re.split("[ ,]+", self.measurements[-1]["bounds"])
            else:
                b = re.split("[ ,]+", self.measurements[0]["bounds"])
            if type(self.measurements[-1]["steps"]) is str:
                s = re.split("[ ,]+", self.measurements[-1]["steps"])
            else:
                s = re.split("[ ,]+", self.measurements[0]["steps"])
            if type(self.measurements[-1]["times"]) is str:
                t = re.split("[ ,]+", self.measurements[-1]["times"])
            else:
                t = re.split("[ ,]+", self.measurements[0]["times"])

            (problem, text) = sanitize_step_scan_parameters(b, s, t)
            if problem is True:
                isok = False
                explanation += f"row {count}:\n" + text
        return isok, explanation

    def skip_row(self, m):
        # all the reasons to skip a line in the spreadsheet
        if type(m["slot"]) is not int:
            return True
        if m["filename"] is None or re.search(r"^\s*$", m["filename"]) is not None:
            return True
        if self.truefalse(m["measure"]) is False:
            return True
        if m["nscans"] is not None and m["nscans"] < 1:
            return True
        return False

    def skip_keyword(self, k):
        """Identify all the keywords that should NOT be captured in the xafs() call."""
        if k in self.experiment or k in self.flags or k in self.motors or k in self.science_metadata:
            return True
        return False

    def make_filename(self, m):
        """Construct a filename with element and edge symbols, if required."""
        fname = m["filename"]
        el = self.measurements[0]["element"]
        ed = self.measurements[0]["edge"]
        if "element" in m:
            el = m["element"]
        if "edge" in m:
            ed = m["edge"]
        if self.append_element.lower() == "element at beginning":
            fname = el + self.joiner + fname
        elif self.append_element.lower() == "element at end":
            fname = fname + self.joiner + el
        elif self.append_element.lower() == "element+edge at beginning":
            fname = el + self.joiner + ed + self.joiner + fname
        elif self.append_element.lower() == "element+edge at end":
            fname = fname + self.joiner + el + self.joiner + ed
        return fname

    # TODO: if time need to be estimated, the code for the function 'conventional_grid' must
    #       be copied from profile collection. Otherwise remove the code for time estimation
    # def estimate_time(self, m, el, ed):
    #     """Approximate the time contribution from the current row"""
    #     if type(m["bounds"]) is str:
    #         b = re.split("[ ,]+", m["bounds"].strip())
    #     else:
    #         b = re.split("[ ,]+", self.measurements[0]["bounds"].strip())
    #     if type(m["steps"]) is str:
    #         s = re.split("[ ,]+", m["steps"].strip())
    #     else:
    #         s = re.split("[ ,]+", self.measurements[0]["steps"].strip())
    #     if type(m["times"]) is str:
    #         t = re.split("[ ,]+", m["times"].strip())
    #     else:
    #         t = re.split("[ ,]+", self.measurements[0]["times"].strip())
    #
    #     b = [float(x) if isfloat(x) else x for x in b]
    #     s = [float(x) if isfloat(x) else x for x in s]
    #     t = [float(x) if isfloat(x) else x for x in t]
    #
    #     (e, t, at, delta) = conventional_grid(
    #         bounds=b, steps=s, times=t, e0=edge_energy(el, ed), element=el, edge=ed, ththth=False
    #     )
    #
    #     if type(m["nscans"]) is int:
    #         nsc = m["nscans"]
    #     else:
    #         nsc = self.measurements[0]["nscans"]
    #     self.totaltime += at * nsc
    #     self.deltatime += delta * delta

    def _find_default_xafs_parameters(self):
        """ Write out the master INI file """

        default = self.measurements[0].copy()
        #  The list of keys that don't go into 'ini' file
        for k in (
            "default",
            "slot",
            "measure",
            "spin",
            "focus",
            "method",
            "samplep",
            "samplex",
            "sampley",
            "slitwidth",
            "detectorx",
        ):
            default.pop(k, None)

        default["url"] = "..."
        default["doi"] = "..."
        default["cif"] = "..."
        default["experimenters"] = self.ws["E1"].value  # top line of xlsx file

        default = self.ini_sanity(default)
        if default is None:
            raise RuntimeError(f"Could not interpret {self.source} as a wheel macro.")

        self.default_xafs_parameters = default

        # Here: 'default' is the dictionary of parameters in 'ini' file. The parameters need
        #   to be passed to the function

    def _gen_plan_list(self):
        raise NotImplementedError("Calling class method that is not implemented")

    def gen_plan_list(self):
        """Write INI file and a BlueSky plan from a spreadsheet.

        Call the subclass' _gen_plan_list to generate the text of the plan.

        """
        self.totaltime, self.deltatime = 0, 0
        self.clear_plan_list()
        self._find_default_xafs_parameters()
        self._gen_plan_list()


class WheelMacroBuilder(BMMMacroBuilder):
    """A class for parsing specially constructed spreadsheets and
    generating macros for measuring XAS on the BMM wheel.

    Examples
    --------
    >>> mb = MacroBuilder()
    >>> mb.spreadsheet('wheel1.xlsx')
    >>> mb.write_macro()
    """

    def _gen_plan_list(self):
        """Write a macro paragraph for each sample described in the
        spreadsheet.  A paragraph consists of line to move to the
        correct wheel slot, a line to change the edge energy (if
        needed), a line to measure the XAFS using the correct set of
        control parameters, and a line to close plot windows after the
        scan.

        Finally, write out the master INI and macro python files.
        """
        element, edge, focus = (None, None, None)
        for m in self.measurements:

            if m["default"] is True:
                element = m["element"]
                edge = m["edge"]
                continue
            if self.skip_row(m) is True:
                continue

            # default element/edge(/focus) values
            for k in ("element", "edge"):
                if m[k] is None:
                    m[k] = self.measurements[0][k]

            # sample and slit movement
            self.add_plan(name="slot", args=[m["slot"]])
            if m["samplex"] is not None:
                self.add_plan(name="mv", args=["xafs_x", m["samplex"]])
            if m["sampley"] is not None:
                self.add_plan(name="mv", args=["xafs_y", m["sampley"]])
            if m["slitwidth"] is not None:
                self.add_plan(name="mv", args=["slits3_hsize", m["slitwidth"]])
            if m["detectorx"] is not None:
                self.add_plan(name="mv", args=["xafs_det", m["detectorx"]])

            # change edge, if needed
            focus = False
            if m["focus"] == "focused":
                focus = True
            if self.do_first_change is True:
                self.add_plan(name="change_edge", args=[m["element"]], kwargs={"edge": m["edge"], "focus": focus})
                self.do_first_change = False
                self.totaltime += 4

            elif m["element"] != element or m["edge"] != edge:  # focus...
                element = m["element"]
                edge = m["edge"]
                self.add_plan(name="change_edge", args=[m["element"]], kwargs={"edge": m["edge"], "focus": focus})
                self.totaltime += 4

            # measure XAFS, then close all plots
            xafs_args = []
            xafs_kwargs = self.default_xafs_parameters.copy()
            command = self.tab + "yield from xafs('%s.ini'" % self.basename
            for k in m.keys():
                # skip cells with macro-building parameters that are not INI parameters
                if self.skip_keyword(k):
                    continue

                # skip element & edge if they are same as default
                elif k in ("element", "edge"):
                    if m[k] == self.measurements[0][k]:
                        continue

                # skip cells with only whitespace
                if type(m[k]) is str and len(m[k].strip()) == 0:
                    m[k] = None

                # if a cell has data, put it in the argument list for xafs()
                if m[k] is not None:
                    if k == "filename":
                        fname = self.make_filename(m)
                        xafs_kwargs["filename"] = fname
                    elif type(m[k]) is int:
                        command += ", %s=%d" % (k, m[k])
                        xafs_kwargs[k] = m[k]
                    elif type(m[k]) is float:
                        xafs_kwargs[k] = m[k]
                    else:
                        xafs_kwargs[k] = str(m[k])

            self.add_plan(name="xafs", args=xafs_args, kwargs=xafs_kwargs)

            # approximate time cost of this sample
            # self.estimate_time(m, element, edge)

        if self.close_shutters:
            self.add_plan(name="shb_close_plan")

    def get_keywords(self, row, defaultline):
        this = {
            "default": defaultline,
            "slot": row[1].value,  # sample location
            "measure": self.truefalse(row[2].value),  # filename and visualization
            "filename": row[3].value,
            "nscans": row[4].value,
            "start": row[5].value,
            "mode": row[6].value,
            # 'e0': row[7].value,
            "element": row[7 + self.offset].value,  # energy range
            "edge": row[8 + self.offset].value,
            "focus": row[9 + self.offset].value,
            "sample": row[10 + self.offset].value,  # scan metadata
            "prep": row[11 + self.offset].value,
            "comment": row[12 + self.offset].value,
            "bounds": row[13 + self.offset].value,  # scan parameters
            "steps": row[14 + self.offset].value,
            "times": row[15 + self.offset].value,
            "samplex": row[16 + self.offset].value,  # other motors
            "sampley": row[17 + self.offset].value,
            "slitwidth": row[18 + self.offset].value,
            "detectorx": row[19 + self.offset].value,
            "snapshots": self.truefalse(row[20 + self.offset].value),  # flags
            "htmlpage": self.truefalse(row[21 + self.offset].value),
            "usbstick": self.truefalse(row[22 + self.offset].value),
            "bothways": self.truefalse(row[23 + self.offset].value),
            "channelcut": self.truefalse(row[24 + self.offset].value),
            "ththth": self.truefalse(row[25 + self.offset].value),
            "url": row[26 + self.offset].value,
            "doi": row[27 + self.offset].value,
            "cif": row[28 + self.offset].value,
        }
        return this


def _unit_test_process_spreadsheet(*, spreadsheet_file):
    """
    Process trivial spreadsheet with plan parameters (for use in unit tests).
    The spreadsheet is expected to contain parameters of 'count' plan in the form:
        name   num   delay
    0   count  5     1
    1   count  6     0.5

    Parameters
    ----------
    spreadsheet_file : file
        readable file object

    Returns
    -------
    plan_list : list(dict)
        Dictionary representing a list of plans extracted from the spreadsheet.
    """
    df = pd.read_excel(spreadsheet_file, index_col=0, engine="openpyxl")
    plan_list = []
    n_rows, _ = df.shape
    for nr in range(n_rows):
        plan = {
            "name": df["name"][nr],
            "args": [["det1", "det2"]],
            "kwargs": {"num": int(df["num"][nr]), "delay": float(df["delay"][nr])},
        }
        plan_list.append(plan)
    return plan_list


def convert_spreadsheet_to_plan_queue(*, spreadsheet_file, instrument_id, data_type, user_name):
    """
    Convert spreadsheet into a list of plans that could be added to the queue.

    Parameters
    ----------
    spreadsheet_file : file
        Readable file object.
    instrument_id : str
        Instrument (beamline) ID, such as ``BMM``. Instrument ID ``__TEST__`` is reserved for
        unit tests.
    data_type : str
        Data type, such as ``excel``. May be used to select proper processing function.
    user_name : str
        User name: may be used as part of plan parameters.

    Returns
    -------
    plan_list : list(dict)
        Dictionary representing a list of plans extracted from the spreadsheet.
    """

    msg_unsupported_data_type = f"Data type '{data_type}' is not supported for the instrument '{instrument_id}'"

    if instrument_id == "BMM":
        if data_type == "excel":
            # pd = pandas.read_excel(
            #     spreadsheet_file, engine="openpyxl", header=None, index_col=None, sheet_name=None
            # )
            # print(f"pd={pd}")
            mb = WheelMacroBuilder(user_name=user_name)
            return mb.process_spreadsheet(spreadsheet_file=spreadsheet_file, energy=True)
        else:
            raise ValueError(msg_unsupported_data_type)
    elif instrument_id == "__TEST__":
        # Special reserved case for use in unit tests.
        if data_type == "excel":
            return _unit_test_process_spreadsheet(spreadsheet_file=spreadsheet_file)
        else:
            raise ValueError(msg_unsupported_data_type)
    elif instrument_id is None:
        raise ValueError(
            "Instrument ID is not set. Set environment variable QSERVER_INSTRUMENT_ID and restart the server"
        )
    else:
        raise ValueError(f"Unsupported instrument: '{instrument_id}'")
